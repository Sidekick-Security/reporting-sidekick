#!/usr/bin/env python3

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

class XLSXGenerator:
    def __init__(self, verbose=False):
        self.verbose = verbose
        self.risk_colors = {
            'Critical': 'FF0000',  # Red
            'High': 'FF8C00',      # Dark Orange  
            'Medium': 'FFD700',    # Gold
            'Low': '90EE90'        # Light Green
        }
        self.risk_priority = {
            'Critical': 4,
            'High': 3,
            'Medium': 2, 
            'Low': 1,
            'Informational': 0
        }
        
    def generate_report(self, findings, output_file, debug_mode=False):
        if not findings:
            raise Exception("No findings to generate report")
            
        # Convert findings to DataFrame
        df = pd.DataFrame(findings)
        
        # Sort by Risk Level, then CVSS score (descending), then group by vulnerability title
        df['risk_priority'] = df['risk'].map(self.risk_priority)
        
        # Convert CVSS score to numeric, handling empty/null values
        df['cvss_numeric'] = pd.to_numeric(df['cvss_score'], errors='coerce').fillna(0)
        
        # Sort by: risk_priority (desc), cvss_numeric (desc), title (asc), host (asc), port (asc)
        df = df.sort_values(
            ['risk_priority', 'cvss_numeric', 'title', 'host', 'port'], 
            ascending=[False, False, True, True, True]
        )
        
        # Clean up temporary columns
        df = df.drop(['risk_priority', 'cvss_numeric'], axis=1)
        
        # Reorder columns for better readability
        if debug_mode:
            # Include debug columns: source, source_file, cvss_vector
            column_order = [
                'risk', 'host', 'hostname', 'port', 'protocol', 'service',
                'title', 'source', 'description', 'solution',
                'cvss_score', 'cve_references', 'source_file',
                'cvss_vector', 'vulnerability_type'
            ]
        else:
            # Standard columns without debug info
            column_order = [
                'risk', 'host', 'hostname', 'port', 'protocol', 'service',
                'title', 'description', 'solution',
                'cvss_score', 'cve_references', 'vulnerability_type'
            ]
        
        # Only include columns that exist in the DataFrame
        available_columns = [col for col in column_order if col in df.columns]
        df = df[available_columns]
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Vulnerability Report"
        
        # Calculate the last column for merging based on number of columns
        from openpyxl.utils import get_column_letter
        last_col = get_column_letter(len(available_columns))
        
        # Set header row based on mode
        if debug_mode:
            # Add title row
            ws.merge_cells(f'A1:{last_col}1')
            title_cell = ws['A1']
            title_cell.value = "Vulnerability Assessment Report"
            title_cell.font = Font(size=16, bold=True)
            title_cell.alignment = Alignment(horizontal='center')
            
            # Add summary row
            ws.merge_cells(f'A2:{last_col}2')
            summary_cell = ws['A2']
            risk_counts = df['risk'].value_counts()
            summary_text = f"Total Findings: {len(df)} | "
            summary_text += " | ".join([f"{risk}: {count}" for risk, count in risk_counts.items()])
            summary_cell.value = summary_text
            summary_cell.font = Font(size=12, bold=True)
            summary_cell.alignment = Alignment(horizontal='center')
            
            header_row = 4
        else:
            # Standard mode - headers start at row 1
            header_row = 1
        
        # Add headers
        if debug_mode:
            headers = [
                'Risk Level', 'Host IP', 'Hostname', 'Port', 'Protocol', 'Service',
                'Vulnerability Title', 'Source', 'Description', 'Remediation',
                'CVSS Score', 'CVE References', 'Source File',
                'CVSS Vector', 'Vulnerability Type'
            ]
        else:
            headers = [
                'Risk Level', 'Host IP', 'Hostname', 'Port', 'Protocol', 'Service',
                'Vulnerability Title', 'Description', 'Remediation',
                'CVSS Score', 'CVE References', 'Vulnerability Type'
            ]
        
        # Trim headers to match available columns
        headers = headers[:len(available_columns)]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Add data rows
        data_start_row = header_row + 1
        for row_idx, (_, row) in enumerate(df.iterrows(), start=data_start_row):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = str(value) if pd.notna(value) else ''
                
                # Apply risk-based coloring to the risk column
                if col_idx == 1 and str(value) in self.risk_colors:  # Risk level column
                    cell.fill = PatternFill(
                        start_color=self.risk_colors[str(value)],
                        end_color=self.risk_colors[str(value)],
                        fill_type='solid'
                    )
                    cell.font = Font(bold=True, color='FFFFFF' if str(value) in ['Critical', 'High'] else '000000')
                
                # Set text wrapping for description and solution columns
                if debug_mode:
                    # In debug mode: Description=9, Solution=10
                    if col_idx in [9, 10]:  
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                else:
                    # In standard mode: Description=8, Solution=9
                    if col_idx in [8, 9]:  
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Auto-adjust column widths
        self._adjust_column_widths(ws, df, debug_mode)
        
        # Freeze panes
        if debug_mode:
            # Freeze panes (freeze first 4 rows only)
            ws.freeze_panes = 'A5'
        else:
            # Freeze panes (freeze first row only)
            ws.freeze_panes = 'A2'
        
        # Create summary sheet
        self._create_summary_sheet(wb, df, debug_mode)
        
        # Save workbook
        wb.save(output_file)
        
        if self.verbose:
            risk_counts = df['risk'].value_counts()
            print(f"Report saved with {len(df)} findings")
            print(f"Risk distribution: {dict(risk_counts)}")
    
    def _adjust_column_widths(self, ws, df, debug_mode=False):
        if debug_mode:
            # Debug mode column widths (includes Source, Source File, CVSS Vector)
            column_widths = {
                1: 12,   # Risk Level
                2: 15,   # Host IP  
                3: 20,   # Hostname
                4: 8,    # Port
                5: 10,   # Protocol
                6: 15,   # Service
                7: 40,   # Title
                8: 10,   # Source
                9: 50,   # Description
                10: 50,  # Solution
                11: 10,  # CVSS Score
                12: 20,  # CVE References
                13: 15,  # Source File
                14: 15,  # CVSS Vector
                15: 20   # Vulnerability Type
            }
        else:
            # Standard column widths (without debug columns)
            column_widths = {
                1: 12,   # Risk Level
                2: 15,   # Host IP  
                3: 20,   # Hostname
                4: 8,    # Port
                5: 10,   # Protocol
                6: 15,   # Service
                7: 40,   # Title
                8: 50,   # Description
                9: 50,   # Solution
                10: 10,  # CVSS Score
                11: 20,  # CVE References
                12: 20   # Vulnerability Type
            }
        
        from openpyxl.utils import get_column_letter
        
        for col_num, width in column_widths.items():
            if col_num <= len(df.columns):
                column_letter = get_column_letter(col_num)
                ws.column_dimensions[column_letter].width = width
    
    def _add_borders(self, ws, max_row, max_col):
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(4, max_row + 1):
            for col in range(1, max_col + 1):
                ws.cell(row=row, column=col).border = thin_border
    
    def _create_summary_sheet(self, wb, df, debug_mode=False):
        summary_ws = wb.create_sheet(title="Summary")
        
        # Title
        summary_ws['A1'] = "Vulnerability Summary Report"
        summary_ws['A1'].font = Font(size=16, bold=True)
        summary_ws.merge_cells('A1:D1')
        
        # Risk distribution
        summary_ws['A3'] = "Risk Distribution:"
        summary_ws['A3'].font = Font(size=14, bold=True)
        
        risk_counts = df['risk'].value_counts().reindex(['Critical', 'High', 'Medium', 'Low'], fill_value=0)
        
        row = 4
        for risk, count in risk_counts.items():
            summary_ws[f'A{row}'] = risk
            summary_ws[f'B{row}'] = count
            summary_ws[f'C{row}'] = f"{count/len(df)*100:.1f}%"
            
            # Apply color coding
            if risk in self.risk_colors:
                for col in ['A', 'B', 'C']:
                    cell = summary_ws[f'{col}{row}']
                    cell.fill = PatternFill(
                        start_color=self.risk_colors[risk],
                        end_color=self.risk_colors[risk],
                        fill_type='solid'
                    )
                    cell.font = Font(bold=True, color='FFFFFF' if risk in ['Critical', 'High'] else '000000')
            row += 1
        
        # Host summary
        summary_ws['A9'] = "Top 10 Hosts by Vulnerability Count:"
        summary_ws['A9'].font = Font(size=14, bold=True)
        
        host_counts = df['host'].value_counts().head(10)
        row = 10
        for host, count in host_counts.items():
            summary_ws[f'A{row}'] = host
            summary_ws[f'B{row}'] = count
            row += 1
        
        # Source distribution (only in debug mode)
        if debug_mode and 'source' in df.columns:
            summary_ws['D3'] = "Finding Sources:"
            summary_ws['D3'].font = Font(size=14, bold=True)
            
            source_counts = df['source'].value_counts()
            row = 4
            for source, count in source_counts.items():
                summary_ws[f'D{row}'] = source
                summary_ws[f'E{row}'] = count
                summary_ws[f'F{row}'] = f"{count/len(df)*100:.1f}%"
                row += 1
        
        # Auto-adjust column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            summary_ws.column_dimensions[col].width = 20